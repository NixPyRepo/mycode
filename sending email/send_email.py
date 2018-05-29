#!/usr/bin/env python3
#Author: Nick Mahoney
'''Description:
	- Script cannot be used on open internet
	- This script with fill out a pre formated time sheet with required information each week
	- It gets the user name and the end of the week (saturday) and fills in the time sheet appropriately
	- Any personal or internal information has been stripped out

	Assumptions:
		Regular 40 hour week. No sick leave or vaction time was used
		Time sheet and script are in the same location

	Future Additions:
		switches to adjust for sick/vacation time
		search for file if script and file are not in the same location
		create/change folders if not in the current month

	Example usage:
		./send_email.py "26_May_Timesheet.xlsx"
'''
import datetime
import openpyxl
import smtplib
import sys
from getpass import getuser
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders

#get the next saturday
def get_saturday():
	
	today = datetime.date.today()
	#Get the numeric value for the day of the week. 0 = Monday 1 = Tuesday ...
	day_index = (today.weekday() + 1) % 7
	
	#Find the date of the upcoming saturday, by adding the difference of todays date
	# plus the number of days until saturday
	sat = (today + datetime.timedelta(6-day_index))
	return sat


def mod_wb(user, name):

	#open workbook to modify
	wb = openpyxl.load_workbook(name)
	#Open the work sheet 
	ws = wb["Timesheet"]
	saturday = get_saturday()
	
	#Set the end of week to upcoming saturday
	ws["G2"] = saturday.strftime("%m/%d/%Y")

	#Set set the userID of the spreadsheet
	ws["G3"] = user

	#Set the date signed to the current date
	ws["D35"] = datetime.date.today().strftime("%m/%d/%Y")
	
	#Create the file name in the form of Month/Date/Year
	fileName = saturday.strftime("%d_%b_Timesheet.xlsx")

	#Save the workbook with new file name
	wb.save(fileName)

def send_mail(user):
	#Change to fit your email domain
	user_email = user + "example.com"	
	
	msg = MIMEMultipart()
	
	saturday = get_saturday()
	endWeek = saturday.strftime("%d %b")
	msg['Subject'] = endWeek + 'Time Sheet'
	msg['From'] = user_email
	#Add appropriate email to send/cc to
	msg['To'] = 'user@example.com'
	msg['cc'] = 'user@example.com'
	
	#Message for the body of the email
	msg.attach(MIMEText("Attached is my time sheet for the week ending in " + endWeek ))

	'''
	- MIMEBase sets the "Content Type" field and usually take "Type/Sub-Type"	
	- MIMEBase has two parameters: Type/Sub-type
	- Aplication/Octet-stream is equivalent to a binary file
	'''
	timesheet = MIMEBase('application', 'octet-stream')
	
	#Get the name of the file to attach in the email
	fileName = saturday.strftime("%d_%b_Timesheet.xlsx")
	#Open and read in the attachment
	timesheet.set_payload(open(fileName, "rb").read())

	'''
	As noted by RFC 821, SMTP only accepts values 0-127, due to 
	some of the bytes in the binary being outside this range  
	encoding the entire attachment in Base64 solves this issue
	'''
	encoders.encode_base64(timesheet)

	#Add the email header declaring the attachment file name
	timesheet.add_header('Content-Disposition', 'attachment; filename='+fileName)
	msg.attach(timesheet)

	#Add the SMTP mail relay
	s = smtplib.SMTP('your-mail-relay.com')
	s.send_message(msg)
	
user = getuser()
#cmd line argument for the file to be opened
name = sys.argv[1]

mod_wb(user, name)
send_mail(user)
